import os
import re
import unicodedata
import psycopg2
import numpy as np
from openpyxl import load_workbook
from pgvector.psycopg2 import register_vector
from sentence_transformers import SentenceTransformer

# =========================
# CONFIGURAÇÕES
# =========================
USUARIO = "postgres"
SENHA = "Fl@qu1nh@"
HOST = "localhost"
PORTA = "5433"
BANCO = "chatbotGI2026"
SCHEMA = "Schemabot"
TABELA = "dados"

CAMINHO_EXCEL = r"C:\Users\sophi\OneDrive\Desktop\Estágio_Obrigatório\ChatBotGI2026\FONTES\PERGUNTAS_RESPOSTAS.xlsx"

ABA_SINONIMOS_APROVADOS = "SINONIMOS_APROVADOS"

MODEL_EMBEDDING = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"

# =========================
# NORMALIZAÇÃO
# =========================
def remover_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )

def limpar_texto(texto: str) -> str:
    texto = texto.lower().strip()
    texto = texto.replace("?", " ")
    texto = texto.replace("!", " ")
    texto = texto.replace(".", " ")
    texto = texto.replace(",", " ")
    texto = texto.replace(";", " ")
    texto = texto.replace(":", " ")
    texto = texto.replace("(", " ")
    texto = texto.replace(")", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()

def expandir_siglas(texto: str) -> str:
    texto = limpar_texto(texto)

    substituicoes = {
        "oque": "o que",
        "oq": "o que",
        "g i": "gestão da informação",
        "gi": "gestão da informação",
        "gestao da informacao": "gestão da informação",
        "curso gi": "curso de gestão da informação",
        "gestor da informacao": "gestor da informação",
        "ciencia da informacao": "ciência da informação",
    }

    texto = f" {texto} "

    for origem, destino in substituicoes.items():
        texto = re.sub(rf"\b{re.escape(origem)}\b", destino, texto)

    return re.sub(r"\s+", " ", texto).strip()

def normalizar(texto: str) -> str:
    texto = expandir_siglas(texto)
    texto = remover_acentos(texto)
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

# =========================
# CONEXÃO
# =========================
print("Carregando modelo de embeddings...")
modelo = SentenceTransformer(MODEL_EMBEDDING)
print("✅ Modelo carregado.")

print("Conectando ao PostgreSQL...")
conn = psycopg2.connect(
    dbname=BANCO,
    user=USUARIO,
    password=SENHA,
    host=HOST,
    port=PORTA
)
register_vector(conn)
cur = conn.cursor()
print("✅ Conexão com PostgreSQL OK.")

# =========================
# FUNÇÕES DE APOIO
# =========================
def pergunta_ja_existe_no_banco(pergunta: str) -> bool:
    pergunta_norm = normalizar(pergunta)

    cur.execute(
        f'''
        SELECT pergunta
        FROM "{SCHEMA}"."{TABELA}"
        '''
    )

    perguntas_bd = cur.fetchall()

    for (pergunta_bd,) in perguntas_bd:
        if normalizar(str(pergunta_bd)) == pergunta_norm:
            return True

    return False

def inserir_pergunta_no_banco(pergunta: str, resposta: str, eixo: str = ""):
    vetor = modelo.encode(
        pergunta,
        normalize_embeddings=True
    ).astype(np.float32)

    cur.execute(
        f'''
        INSERT INTO "{SCHEMA}"."{TABELA}"
        (pergunta, resposta, eixo, embedding)
        VALUES (%s, %s, %s, %s)
        ''',
        (pergunta, resposta, eixo, vetor)
    )

    conn.commit()

def carregar_base_excel():
    if not os.path.exists(CAMINHO_EXCEL):
        raise FileNotFoundError(f"Arquivo não encontrado: {CAMINHO_EXCEL}")

    wb = load_workbook(CAMINHO_EXCEL)

    aba_ativa = wb[wb.sheetnames[0]]

    registros = []

    # tenta detectar colunas da aba principal
    cabecalho = [str(c.value).strip().lower() if c.value else "" for c in aba_ativa[1]]

    mapa_colunas = {}
    for i, nome in enumerate(cabecalho):
        mapa_colunas[nome] = i

    # aceita formatos comuns
    idx_pergunta = None
    idx_resposta = None
    idx_eixo = None
    idx_tipo = None

    for nome, idx in mapa_colunas.items():
        if nome in ["pergunta", "questao"]:
            idx_pergunta = idx
        elif nome in ["resposta", "resposta_curta", "resposta longa", "resposta_longa"]:
            if idx_resposta is None:
                idx_resposta = idx
        elif nome in ["eixo", "tipo"]:
            if idx_eixo is None:
                idx_eixo = idx
        elif nome == "tipo":
            idx_tipo = idx

    if idx_pergunta is None or idx_resposta is None:
        raise ValueError("Não encontrei colunas de PERGUNTA e RESPOSTA na primeira aba do Excel.")

    for row in aba_ativa.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        pergunta = row[idx_pergunta] if idx_pergunta < len(row) else None
        resposta = row[idx_resposta] if idx_resposta < len(row) else None

        if not pergunta or not resposta:
            continue

        eixo = ""
        if idx_eixo is not None and idx_eixo < len(row) and row[idx_eixo] is not None:
            eixo = str(row[idx_eixo]).strip()

        registros.append({
            "pergunta": str(pergunta).strip(),
            "resposta": str(resposta).strip(),
            "eixo": eixo
        })

    # carregar sinônimos aprovados
    if ABA_SINONIMOS_APROVADOS in wb.sheetnames:
        ws_sinonimos = wb[ABA_SINONIMOS_APROVADOS]

        for row in ws_sinonimos.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            # formato:
            # data, pergunta_usuario, pergunta_normalizada, pergunta_canonica, confianca
            if len(row) < 5:
                continue

            _, pergunta_usuario, _, pergunta_canonica, _ = row

            if not pergunta_usuario or not pergunta_canonica:
                continue

            # localizar resposta da pergunta canônica
            resposta_canonica = None
            eixo_canonico = ""

            for item in registros:
                if normalizar(item["pergunta"]) == normalizar(str(pergunta_canonica)):
                    resposta_canonica = item["resposta"]
                    eixo_canonico = item["eixo"]
                    break

            # se não achou na aba principal, tenta no banco
            if resposta_canonica is None:
                cur.execute(
                    f'''
                    SELECT resposta, eixo
                    FROM "{SCHEMA}"."{TABELA}"
                    WHERE lower(pergunta) = lower(%s)
                    LIMIT 1
                    ''',
                    (str(pergunta_canonica).strip(),)
                )
                achado = cur.fetchone()
                if achado:
                    resposta_canonica = achado[0]
                    eixo_canonico = achado[1] if achado[1] else ""

            if resposta_canonica:
                registros.append({
                    "pergunta": str(pergunta_usuario).strip(),
                    "resposta": str(resposta_canonica).strip(),
                    "eixo": str(eixo_canonico).strip() if eixo_canonico else ""
                })

    return registros

# =========================
# SINCRONIZAÇÃO
# =========================
def sincronizar():
    registros = carregar_base_excel()

    total = len(registros)
    inseridos = 0
    ignorados = 0

    print(f"📄 Total de registros lidos do Excel: {total}")

    for item in registros:
        pergunta = item["pergunta"]
        resposta = item["resposta"]
        eixo = item["eixo"]

        if pergunta_ja_existe_no_banco(pergunta):
            ignorados += 1
            print(f"⏭️ Já existe no banco: {pergunta}")
            continue

        inserir_pergunta_no_banco(pergunta, resposta, eixo)
        inseridos += 1
        print(f"✅ Inserido no banco: {pergunta}")

    print("\n=== RESUMO DA SINCRONIZAÇÃO ===")
    print(f"Total lido do Excel: {total}")
    print(f"Inseridos no banco: {inseridos}")
    print(f"Ignorados (já existiam): {ignorados}")

# =========================
# EXECUÇÃO
# =========================
try:
    sincronizar()
except Exception as e:
    print(f"❌ Erro durante a sincronização: {e}")
finally:
    cur.close()
    conn.close()