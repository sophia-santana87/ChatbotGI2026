import os
import re
import unicodedata
from datetime import datetime

import psycopg2
import numpy as np
from openpyxl import load_workbook, Workbook
from pgvector.psycopg2 import register_vector
from sentence_transformers import SentenceTransformer

# =========================
# CONFIGURAÇÕES
# =========================
USUARIO = "postgres"
SENHA = "Fl@qu1nh@"
HOST = "localhost"
PORTA = "5433"
BANCO = "chatbotGI2026"
SCHEMA = "Schemabot"
TABELA = "dados"

CAMINHO_EXCEL = r"C:\Users\sophi\OneDrive\Desktop\Estágio_Obrigatório\ChatBotGI2026\FONTES\PERGUNTAS_RESPOSTAS.xlsx"

MODEL_EMBEDDING = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"

TOP_K_VETOR = 5
TOP_K_FINAL = 5

PESO_TEXTO = 0.65
PESO_VETOR = 0.35

LIMIAR_RESPOSTA_DIRETA = 45
LIMIAR_MINIMO = 25

LIMIAR_SALVAR_APROVADO = 60
LIMIAR_SALVAR_PENDENTE = 35

ABA_SINONIMOS_APROVADOS = "SINONIMOS_APROVADOS"
ABA_SINONIMOS_PENDENTES = "SINONIMOS_PENDENTES"

# =========================
# MODELO
# =========================
print("Carregando modelo de embeddings...")
modelo = SentenceTransformer(MODEL_EMBEDDING)
print("✅ Modelo carregado.")

# =========================
# CONEXÃO BANCO
# =========================
print("Conectando ao PostgreSQL...")
conn = psycopg2.connect(
    dbname=BANCO,
    user=USUARIO,
    password=SENHA,
    host=HOST,
    port=PORTA
)

register_vector(conn)
cur = conn.cursor()

print("✅ Conexão com PostgreSQL OK.")

# =========================
# NORMALIZAÇÃO
# =========================
def remover_acentos(texto):

    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def limpar_texto(texto):

    texto = texto.lower()

    texto = texto.replace("?", " ")
    texto = texto.replace("!", " ")
    texto = texto.replace(".", " ")
    texto = texto.replace(",", " ")

    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def expandir_siglas(texto):

    texto = limpar_texto(texto)

    mapa = {
        "oque": "o que",
        "oq": "o que",
        "gi": "gestão da informação",
        "g i": "gestão da informação",
        "gestao da informacao": "gestão da informação"
    }

    texto = f" {texto} "

    for origem, destino in mapa.items():

        texto = re.sub(rf"\b{origem}\b", destino, texto)

    return texto.strip()


def normalizar(texto):

    texto = expandir_siglas(texto)
    texto = remover_acentos(texto)
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def tokenizar(texto):

    texto = normalizar(texto)

    return [t for t in texto.split() if len(t) > 1]

# =========================
# PLANILHA
# =========================
def garantir_planilha():

    pasta = os.path.dirname(CAMINHO_EXCEL)

    if not os.path.exists(pasta):
        os.makedirs(pasta)

    if not os.path.exists(CAMINHO_EXCEL):

        wb = Workbook()

        ws = wb.active
        ws.title = "BASE"

        wb.save(CAMINHO_EXCEL)

    wb = load_workbook(CAMINHO_EXCEL)

    if ABA_SINONIMOS_APROVADOS not in wb.sheetnames:

        ws = wb.create_sheet(ABA_SINONIMOS_APROVADOS)

        ws.append([
            "data",
            "pergunta_usuario",
            "pergunta_normalizada",
            "pergunta_canonica",
            "confianca"
        ])

    if ABA_SINONIMOS_PENDENTES not in wb.sheetnames:

        ws = wb.create_sheet(ABA_SINONIMOS_PENDENTES)

        ws.append([
            "data",
            "pergunta_usuario",
            "pergunta_normalizada",
            "pergunta_canonica",
            "confianca"
        ])

    wb.save(CAMINHO_EXCEL)


garantir_planilha()

# =========================
# CARREGAR BASE
# =========================
def carregar_base():

    cur.execute(
        f"""
        SELECT id, pergunta, resposta, eixo
        FROM "{SCHEMA}"."{TABELA}"
        """
    )

    return cur.fetchall()


BASE = carregar_base()

# =========================
# BUSCA TEXTUAL
# =========================
def score_textual(pergunta_usuario, pergunta_bd):

    tokens_u = set(tokenizar(pergunta_usuario))
    tokens_bd = set(tokenizar(pergunta_bd))

    inter = len(tokens_u & tokens_bd)

    base = max(len(tokens_u), 1)

    return (inter / base) * 100

# =========================
# BUSCA VETORIAL
# =========================
def buscar_vetorial(pergunta):

    pergunta = expandir_siglas(pergunta)

    vetor = modelo.encode(
        pergunta,
        normalize_embeddings=True
    ).astype(np.float32)

    cur.execute(
        f"""
        SELECT
            id,
            pergunta,
            resposta,
            eixo,
            embedding <=> %s::vector AS distancia
        FROM "{SCHEMA}"."{TABELA}"
        ORDER BY embedding <=> %s::vector
        LIMIT %s
        """,
        (vetor, vetor, TOP_K_VETOR)
    )

    rows = cur.fetchall()

    resultados = []

    for r in rows:

        similaridade = (1 - float(r[4])) * 100

        resultados.append({
            "id": r[0],
            "pergunta": r[1],
            "resposta": r[2],
            "eixo": r[3],
            "score_vetor": similaridade,
            "score_texto": 0
        })

    return resultados

# =========================
# INSERIR SINÔNIMO NO BANCO
# =========================
def inserir_sinonimo_no_banco(pergunta_usuario, pergunta_canonica):

    for id_reg, pergunta_bd, resposta_bd, eixo_bd in BASE:

        if normalizar(pergunta_bd) == normalizar(pergunta_canonica):

            resposta = resposta_bd
            eixo = eixo_bd
            break
    else:
        return

    vetor = modelo.encode(
        pergunta_usuario,
        normalize_embeddings=True
    ).astype(np.float32)

    cur.execute(
        f"""
        INSERT INTO "{SCHEMA}"."{TABELA}"
        (pergunta, resposta, eixo, embedding)
        VALUES (%s,%s,%s,%s)
        """,
        (pergunta_usuario, resposta, eixo, vetor)
    )

    conn.commit()

    print("📚 Novo sinônimo adicionado ao banco.")

# =========================
# SALVAR SINÔNIMO
# =========================
def salvar_excel(aba, pergunta_usuario, pergunta_canonica, confianca):

    wb = load_workbook(CAMINHO_EXCEL)

    ws = wb[aba]

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        pergunta_usuario,
        normalizar(pergunta_usuario),
        pergunta_canonica,
        round(confianca,2)
    ])

    wb.save(CAMINHO_EXCEL)

# =========================
# APRENDIZADO
# =========================
def aprender(pergunta_usuario, melhor):

    pergunta_canonica = melhor["pergunta"]

    confianca = melhor["score_final"]

    if confianca >= LIMIAR_SALVAR_APROVADO:

        salvar_excel(
            ABA_SINONIMOS_APROVADOS,
            pergunta_usuario,
            pergunta_canonica,
            confianca
        )

        inserir_sinonimo_no_banco(
            pergunta_usuario,
            pergunta_canonica
        )

    elif confianca >= LIMIAR_SALVAR_PENDENTE:

        salvar_excel(
            ABA_SINONIMOS_PENDENTES,
            pergunta_usuario,
            pergunta_canonica,
            confianca
        )

# =========================
# RESPOSTA
# =========================
def responder(pergunta):

    resultados_vetor = buscar_vetorial(pergunta)

    resultados = []

    for r in resultados_vetor:

        st = score_textual(pergunta, r["pergunta"])

        r["score_texto"] = st

        r["score_final"] = (
            st * PESO_TEXTO +
            r["score_vetor"] * PESO_VETOR
        )

        resultados.append(r)

    resultados.sort(
        key=lambda x: x["score_final"],
        reverse=True
    )

    melhor = resultados[0]

    confianca = melhor["score_final"]

    if confianca >= LIMIAR_RESPOSTA_DIRETA:

        aprender(pergunta, melhor)

        print("\n🤖 Assistente:", melhor["resposta"])
        print(f"✅ Confiança {round(confianca,1)}%")

        return

    if confianca >= LIMIAR_MINIMO:

        aprender(pergunta, melhor)

        print("\n🤖 Assistente:", melhor["resposta"])
        print(f"⚠️ Confiança {round(confianca,1)}%")

        return

    print("\n🤖 Não encontrei essa informação.")

# =========================
# LOOP
# =========================
print("\n=== Chatbot GI UFPR iniciado ===")

while True:

    pergunta = input("\nVocê: ")

    if pergunta.lower() in ["sair","exit","quit"]:
        break

    responder(pergunta)